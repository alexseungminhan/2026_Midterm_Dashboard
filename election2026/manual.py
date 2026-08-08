"""manual.py — first-class manual data ingestion (polls, ad spend, ratings).

No free API exists for these sources, so user-maintained spreadsheets in
data/manual/ are treated as authoritative inputs — validated on load, cached
alongside API pulls, and indistinguishable downstream from an API source.

    python -m election2026 make-templates    # writes blank .xlsx templates

Validation failures produce actionable messages naming the offending row —
never a stack trace.
"""

from __future__ import annotations

import math
import os
from datetime import date, datetime
from typing import Optional

from . import config

TEMPLATES = {
    "polls_template.xlsx": ["race_id", "pollster", "date", "sample_size",
                            "margin_dem", "matchup", "weight", "notes"],
    "ratings_template.xlsx": ["race_id", "rating", "as_of", "source",
                              "notes"],
    # Track B manual sources (see track_b/adapters.py).
    "primary_turnout_template.xlsx": [
        "state", "cycle", "dem_votes", "rep_votes", "contested_dem",
        "contested_rep", "certified_date", "source_url", "notes"],
    "party_registration_template.xlsx": [
        "state", "report_date", "dem_registered", "rep_registered",
        "source_url", "notes"],
    # Relative form of the same quantity — see load_primary_turnout_change.
    "primary_turnout_change_template.xlsx": [
        "state", "cycle", "prior_cycle", "dem_pct_of_prior",
        "rep_pct_of_prior", "contest", "source_url", "notes"],
    # adspend_template.xlsx was REMOVED 2026-07-30: commercial ad-tracking data
    # is paywalled and the free alternative (FCC political files) is scanned
    # PDFs whose parsing cost exceeds the signal. Do not re-add.
}

# Rows land in data/manual/<name>.xlsx (or .csv) — the "_template" suffix is
# only for the blank files make-templates writes.
FILES = {
    "polls": "polls",
    "ratings": "ratings",
    "primary_turnout": "primary_turnout",
    "primary_turnout_change": "primary_turnout_change",
    "party_registration": "party_registration",
}


class ManualDataError(ValueError):
    """A manual spreadsheet failed validation. Message names the bad rows."""


# ---------------------------------------------------------------------------
# Matchup confidence
# ---------------------------------------------------------------------------
# A poll of a matchup that will not be on the ballot still carries information
# about the partisan environment of the seat — it is just weaker evidence than
# a poll of the actual nominees. Dropping those rows outright (the behaviour
# before 2026-07-30) left Michigan, New Hampshire and Maine with NO polls at
# all, so their p_consensus came from betting markets alone. Reading them as a
# generic "would you back the Democrat or the Republican" signal at a reduced
# weight is strictly more information than discarding them.
#
# The weight multiplies the recency x sqrt(n) weight in aggregate_polls.
MATCHUP_WEIGHTS = {
    # Both nominees settled and polled by name.
    "confirmed": 1.0,
    # Named-party generic ballot in the seat ("Dem." vs "Rep."). No candidate
    # quality, but it prices the right event and needs no nominee.
    "generic_ballot": 0.60,
    # A named matchup before the primary settled it — one or both names may
    # not reach the ballot.
    "hypothetical": 0.35,
    # A named matchup where a polled candidate has since withdrawn.
    "withdrawn": 0.15,
}
MATCHUP_KINDS = tuple(MATCHUP_WEIGHTS)


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------

def make_templates(directory: Optional[str] = None) -> list:
    """Write blank, correctly-headed templates into data/manual/."""
    from openpyxl import Workbook

    directory = directory or config.MANUAL_DIR
    os.makedirs(directory, exist_ok=True)
    written = []
    for fname, headers in TEMPLATES.items():
        wb = Workbook()
        ws = wb.active
        ws.title = fname.replace("_template.xlsx", "")
        ws.append(headers)
        # One illustrative row so column formats are obvious.
        example = {
            "polls_template.xlsx":
                ["sen-ga", "ExamplePoll/UGA", "2026-07-20", 800, 2.5,
                 "confirmed", 1.0,
                 "margin_dem = D% - R%, positive = Dem ahead; matchup ∈ "
                 + "/".join(MATCHUP_KINDS)],
            "ratings_template.xlsx":
                ["sen-ga", "Lean D", "2026-07-20", "Cook",
                 "rating ∈ Tossup/Lean D/Lean R/Likely D/Likely R/Safe D/Safe R"],
            "primary_turnout_template.xlsx":
                ["GA", 2026, 742000, 1180000, True, True, "2026-06-16",
                 "https://sos.ga.gov/...",
                 "votes cast in each party's statewide primary; add prior "
                 "midterm cycles (2018/2022) as their own rows for the "
                 "state baseline"],
            "primary_turnout_change_template.xlsx":
                ["IA", 2026, 2022, 122, 108, "governor",
                 "https://sos.iowa.gov/elections/...",
                 "each party's 2026 primary turnout as a PERCENT of its own "
                 "2022 turnout; 100 = unchanged"],
            "party_registration_template.xlsx":
                ["NC", "2026-07-25", 2380000, 2290000,
                 "https://www.ncsbe.gov/results-data/voter-registration-data",
                 "registered voters by party as of report_date; only AK, ME, "
                 "NH, NC, IA, NE have party registration"],
        }[fname]
        ws.append(example)
        path = os.path.join(directory, fname)
        wb.save(path)
        written.append(path)
    return written


# ---------------------------------------------------------------------------
# Loading (.xlsx via openpyxl, .csv via stdlib)
# ---------------------------------------------------------------------------

def _find_file(name: str, directory: str) -> Optional[str]:
    for ext in (".xlsx", ".csv"):
        path = os.path.join(directory, name + ext)
        if os.path.exists(path):
            return path
    return None


def _read_rows(path: str) -> list:
    """Return a list of dicts keyed by the header row."""
    if path.endswith(".csv"):
        import csv
        with open(path, encoding="utf-8-sig", newline="") as fh:
            return [dict(row) for row in csv.DictReader(fh)]
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    out = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        out.append({h: v for h, v in zip(headers, values) if h})
    return out


def _parse_date(value, field: str, row_no: int, fname: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        raise ManualDataError(
            "%s row %d: %s=%r is not a date — use YYYY-MM-DD"
            % (fname, row_no, field, value))


def _parse_float(value, field: str, row_no: int, fname: str,
                 lo=None, hi=None) -> float:
    try:
        x = float(value)
    except Exception:
        raise ManualDataError("%s row %d: %s=%r is not a number"
                              % (fname, row_no, field, value))
    if lo is not None and x < lo or hi is not None and x > hi:
        raise ManualDataError(
            "%s row %d: %s=%s outside the plausible range [%s, %s]"
            % (fname, row_no, field, x, lo, hi))
    return x


_TRUE = {"true", "t", "yes", "y", "1", "예", "o"}
_FALSE = {"false", "f", "no", "n", "0", "아니오", "x"}


def _parse_bool(value, field: str, row_no: int, fname: str) -> Optional[bool]:
    """Tri-state: None when the cell is blank, else a strict boolean."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    token = str(value).strip().lower()
    if token in _TRUE:
        return True
    if token in _FALSE:
        return False
    raise ManualDataError(
        "%s row %d: %s=%r is not a yes/no value (use TRUE or FALSE)"
        % (fname, row_no, field, value))


def _check_race_id(value, row_no: int, fname: str) -> str:
    rid = str(value or "").strip()
    if rid not in config.RACES:
        known = ", ".join(sorted(config.RACES)[:6])
        raise ManualDataError(
            "%s row %d: unknown race_id %r — must be one of the monitored "
            "races (e.g. %s, ...); see data/reference/races.json"
            % (fname, row_no, rid, known))
    return rid


# ---------------------------------------------------------------------------
# Polls (Track A manual source)
# ---------------------------------------------------------------------------

def load_polls(directory: Optional[str] = None) -> dict:
    """{race_id: [{pollster, date, sample_size, margin_dem, matchup, weight}]}.

    Returns {} when no polls file exists (source unavailable, not an error).

    `matchup` names the matchup-confidence class (MATCHUP_WEIGHTS) and
    `weight` is the multiplier actually applied. Either may be omitted: a row
    with neither is treated as `confirmed` at 1.0, which is what every row in
    a pre-2026-07-30 sheet was.
    """
    directory = directory or config.MANUAL_DIR
    path = _find_file(FILES["polls"], directory)
    if path is None:
        return {}
    fname = os.path.basename(path)
    out: dict = {}
    for i, row in enumerate(_read_rows(path), start=2):
        rid = _check_race_id(row.get("race_id"), i, fname)
        when = _parse_date(row.get("date"), "date", i, fname)
        n = row.get("sample_size")
        n = int(_parse_float(n, "sample_size", i, fname, lo=50, hi=100000)) \
            if n not in (None, "") else None
        margin = _parse_float(row.get("margin_dem"), "margin_dem", i, fname,
                              lo=-60, hi=60)
        matchup = str(row.get("matchup") or "").strip() or "confirmed"
        if matchup not in MATCHUP_WEIGHTS:
            raise ManualDataError(
                "%s row %d: matchup=%r must be one of %s"
                % (fname, i, row.get("matchup"), "/".join(MATCHUP_KINDS)))
        weight = row.get("weight")
        weight = (_parse_float(weight, "weight", i, fname, lo=0.0, hi=1.0)
                  if weight not in (None, "") else MATCHUP_WEIGHTS[matchup])
        out.setdefault(rid, []).append({
            "pollster": str(row.get("pollster") or "").strip() or None,
            "date": when.isoformat(),
            "sample_size": n,
            "margin_dem": margin,
            "matchup": matchup,
            "weight": weight,
        })
    return out


def aggregate_polls(polls: list, as_of: Optional[date] = None,
                    halflife_days: Optional[float] = None) -> Optional[float]:
    """Aggregate a race's polls into one margin (D+points).

    Weight = recency (exponential decay, configurable half-life) × sqrt(n)
    (sample size) × matchup confidence (MATCHUP_WEIGHTS — a poll of a matchup
    that may not reach the ballot counts, but counts less). Returns None when
    the list is empty or every row carries zero weight.
    """
    if not polls:
        return None
    as_of = as_of or date.today()
    halflife = halflife_days or config.TRACK_A["poll_halflife_days"]
    num = den = 0.0
    for p in polls:
        age = (as_of - date.fromisoformat(p["date"])).days
        w_recency = 0.5 ** (max(age, 0) / halflife)
        w_size = math.sqrt(p["sample_size"]) if p.get("sample_size") else math.sqrt(500)
        w_matchup = p.get("weight")
        if w_matchup is None:
            w_matchup = MATCHUP_WEIGHTS.get(p.get("matchup", "confirmed"), 1.0)
        w = w_recency * w_size * float(w_matchup)
        num += w * p["margin_dem"]
        den += w
    return num / den if den > 0 else None


def poll_confidence(polls: list, as_of: Optional[date] = None,
                    halflife_days: Optional[float] = None) -> float:
    """Share of a race's aggregate poll weight coming from settled matchups.

    1.0 = every poll is of the actual nominees; 0.35 = the average is carried
    by pre-primary hypotheticals. Reported alongside the margin so a reader
    can see that Michigan's poll channel is not the same kind of evidence as
    Georgia's, rather than having both arrive as a bare probability.
    """
    if not polls:
        return 0.0
    as_of = as_of or date.today()
    halflife = halflife_days or config.TRACK_A["poll_halflife_days"]
    num = den = 0.0
    for p in polls:
        age = (as_of - date.fromisoformat(p["date"])).days
        base = (0.5 ** (max(age, 0) / halflife)) * (
            math.sqrt(p["sample_size"]) if p.get("sample_size")
            else math.sqrt(500))
        w_matchup = p.get("weight")
        if w_matchup is None:
            w_matchup = MATCHUP_WEIGHTS.get(p.get("matchup", "confirmed"), 1.0)
        num += base * float(w_matchup)
        den += base
    return round(num / den, 4) if den > 0 else 0.0


# ---------------------------------------------------------------------------
# Ad spend — REMOVED 2026-07-30. load_adspend() and its template are gone
# because commercial ad-tracking data is paywalled and the free alternative
# (FCC political files) is largely scanned PDFs whose parsing cost outweighs
# the signal. Do not casually re-add.
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Primary turnout by party (Track B manual source — see adapters.py)
# ---------------------------------------------------------------------------

# States whose primaries cannot be partitioned by party electorate. Alaska
# runs a top-four jungle primary with ranked-choice voting: every voter gets
# the same ballot, so "votes cast in the Democratic primary" does not exist as
# a quantity. Coercing one out of the candidate totals would invent a number.
NON_PARTISAN_PRIMARY_STATES = {
    "AK": "top-four jungle primary with ranked-choice voting — no "
          "party-partitioned primary electorate exists",
}

# Sanity band on dem_votes / rep_votes. Real midterm primary ratios in these
# states have run roughly 0.3–3.0; anything outside is a transcription error
# (a digit dropped, or one party's number pasted twice), not a landslide.
PRIMARY_RATIO_RANGE = (0.05, 20.0)


def load_primary_turnout(directory: Optional[str] = None) -> dict:
    """{state: {cycle: {dem_votes, rep_votes, contested_*, certified_date}}}.

    Validation is deliberately loud: this file is typed by hand once per
    cycle from Secretary of State certifications, so a bad row must name
    itself rather than propagate a wrong ratio into every later run.
    """
    directory = directory or config.MANUAL_DIR
    path = _find_file(FILES["primary_turnout"], directory)
    if path is None:
        return {}
    fname = os.path.basename(path)
    known_states = {m["state"] for m in config.RACES.values()}
    out: dict = {}
    for i, row in enumerate(_read_rows(path), start=2):
        state = str(row.get("state") or "").strip().upper()
        if len(state) != 2 or not state.isalpha():
            raise ManualDataError(
                "%s row %d: state=%r is not a two-letter code"
                % (fname, i, row.get("state")))
        if state not in known_states:
            raise ManualDataError(
                "%s row %d: state=%r is not a monitored state (%s)"
                % (fname, i, state, ", ".join(sorted(known_states))))
        cycle = row.get("cycle")
        if cycle in (None, ""):
            raise ManualDataError("%s row %d: cycle is missing — every row "
                                  "needs the election year" % (fname, i))
        cycle = int(_parse_float(cycle, "cycle", i, fname, lo=1990, hi=2026))
        dem = _parse_float(row.get("dem_votes"), "dem_votes", i, fname, lo=0)
        rep = _parse_float(row.get("rep_votes"), "rep_votes", i, fname, lo=0)
        if state in NON_PARTISAN_PRIMARY_STATES:
            raise ManualDataError(
                "%s row %d: %s has no party-partitioned primary (%s) — remove "
                "the row; the variable is emitted as structurally "
                "unavailable, not as a number"
                % (fname, i, state, NON_PARTISAN_PRIMARY_STATES[state]))
        if rep <= 0 or dem <= 0:
            raise ManualDataError(
                "%s row %d: %s %d has dem_votes=%s rep_votes=%s — a zero makes "
                "the ratio undefined; leave the row out instead"
                % (fname, i, state, cycle, dem, rep))
        ratio = dem / rep
        lo, hi = PRIMARY_RATIO_RANGE
        if not lo <= ratio <= hi:
            raise ManualDataError(
                "%s row %d: %s %d gives dem/rep = %.2f, outside the plausible "
                "range [%s, %s] — check the vote totals"
                % (fname, i, state, cycle, ratio, lo, hi))
        entry = {
            "dem_votes": dem, "rep_votes": rep, "ratio": ratio,
            "contested_dem": _parse_bool(row.get("contested_dem"),
                                         "contested_dem", i, fname),
            "contested_rep": _parse_bool(row.get("contested_rep"),
                                         "contested_rep", i, fname),
            "certified_date": (
                _parse_date(row["certified_date"], "certified_date", i,
                            fname).isoformat()
                if row.get("certified_date") not in (None, "") else None),
            "source_url": str(row.get("source_url") or "").strip() or None,
        }
        by_cycle = out.setdefault(state, {})
        if cycle in by_cycle:
            raise ManualDataError(
                "%s row %d: %s %d appears twice — one row per state per cycle"
                % (fname, i, state, cycle))
        by_cycle[cycle] = entry
    return out


# ---------------------------------------------------------------------------
# Primary turnout, RELATIVE form (Track B manual source — see adapters.py)
# ---------------------------------------------------------------------------

# Published turnout comparisons report each party's primary turnout as a
# percentage of its own turnout in the previous midterm, rather than as vote
# counts. That form is not a downgrade — it is the SAME quantity the
# vote-count path computes, one algebraic step further along:
#
#     log( (D26/D22) / (R26/R22) )  ==  log(D26/R26) - log(D22/R22)
#
# which is exactly "how far this cycle's D/R turnout ratio moved from last
# cycle's", the deviation load_primary_turnout builds from levels. It is
# reachable from a two-number-per-state table, whereas the vote-count path
# needs four numbers per state and at least two prior cycles.
#
# Verified 2026-08-02 on Iowa against the Secretary of State's own certified
# turnout PDFs: SOS gives D 122.7% / R 109.2%, the published comparison gives
# 122% / 108%. The two agree to rounding, so the relative table is not a
# weaker source, only a more compact one.

PRIMARY_CHANGE_RANGE = (5.0, 1000.0)   # percent-of-prior sanity band


def load_primary_turnout_change(directory: Optional[str] = None) -> dict:
    """{state: {dem_pct, rep_pct, cycle, prior_cycle, contest, source_url}}.

    Each percentage is that party's turnout in `cycle` as a percent of its own
    turnout in `prior_cycle` — 100 means unchanged. Returns {} when the file
    is absent, which is a missing source and not an error.
    """
    directory = directory or config.MANUAL_DIR
    path = _find_file(FILES["primary_turnout_change"], directory)
    if path is None:
        return {}
    fname = os.path.basename(path)
    out: dict = {}
    for i, row in enumerate(_read_rows(path), start=2):
        state = str(row.get("state") or "").strip().upper()
        if len(state) != 2 or not state.isalpha():
            raise ManualDataError(
                "%s row %d: state=%r is not a two-letter code"
                % (fname, i, row.get("state")))
        if state in NON_PARTISAN_PRIMARY_STATES:
            raise ManualDataError(
                "%s row %d: %s has no party-partitioned primary (%s) — remove "
                "the row" % (fname, i, state,
                             NON_PARTISAN_PRIMARY_STATES[state]))
        lo, hi = PRIMARY_CHANGE_RANGE
        dem = _parse_float(row.get("dem_pct_of_prior"), "dem_pct_of_prior",
                           i, fname, lo=lo, hi=hi)
        rep = _parse_float(row.get("rep_pct_of_prior"), "rep_pct_of_prior",
                           i, fname, lo=lo, hi=hi)
        cycle = int(_parse_float(row.get("cycle") or 2026, "cycle", i, fname,
                                 lo=1990, hi=2026))
        prior = int(_parse_float(row.get("prior_cycle") or 2022,
                                 "prior_cycle", i, fname, lo=1990, hi=2026))
        if prior >= cycle:
            raise ManualDataError(
                "%s row %d: prior_cycle %d must be earlier than cycle %d"
                % (fname, i, prior, cycle))
        if state in out:
            raise ManualDataError(
                "%s row %d: %s appears twice — one row per state"
                % (fname, i, state))
        out[state] = {
            "dem_pct": dem, "rep_pct": rep, "cycle": cycle,
            "prior_cycle": prior,
            "contest": str(row.get("contest") or "").strip() or None,
            "source_url": str(row.get("source_url") or "").strip() or None,
        }
    return out


# ---------------------------------------------------------------------------
# Party registration (Track B manual source — see adapters.py)
# ---------------------------------------------------------------------------

# Only these states record a voter's party at registration. The other four
# monitored states (MI, OH, GA, TX) have no party registration AT ALL — that
# is a permanent structural fact, not a gap in our collection, and the
# pipeline must report the two differently.
PARTY_REGISTRATION_STATES = {"AK", "ME", "NH", "NC", "IA", "NE"}


def load_party_registration(directory: Optional[str] = None) -> dict:
    """{state: [{report_date, dem_registered, rep_registered}]} oldest first.

    Report dates are kept per observation because states publish on different
    cadences (North Carolina weekly, others monthly or quarterly), and the
    net-change deviation is only meaningful over comparable windows.
    """
    directory = directory or config.MANUAL_DIR
    path = _find_file(FILES["party_registration"], directory)
    if path is None:
        return {}
    fname = os.path.basename(path)
    out: dict = {}
    for i, row in enumerate(_read_rows(path), start=2):
        state = str(row.get("state") or "").strip().upper()
        if state not in PARTY_REGISTRATION_STATES:
            raise ManualDataError(
                "%s row %d: state=%r has no party registration — only %s do. "
                "Remove the row; the variable is emitted as structurally "
                "unavailable for the others."
                % (fname, i, row.get("state"),
                   ", ".join(sorted(PARTY_REGISTRATION_STATES))))
        when = _parse_date(row.get("report_date"), "report_date", i, fname)
        dem = _parse_float(row.get("dem_registered"), "dem_registered", i,
                           fname, lo=0, hi=5e7)
        rep = _parse_float(row.get("rep_registered"), "rep_registered", i,
                           fname, lo=0, hi=5e7)
        out.setdefault(state, []).append({
            "report_date": when.isoformat(),
            "dem_registered": dem,
            "rep_registered": rep,
            "source_url": str(row.get("source_url") or "").strip() or None,
        })
    for rows in out.values():
        rows.sort(key=lambda r: r["report_date"])
        seen = set()
        for r in rows:
            if r["report_date"] in seen:
                raise ManualDataError(
                    "%s: duplicate report_date %s — one row per state per "
                    "report" % (fname, r["report_date"]))
            seen.add(r["report_date"])
    return out


# ---------------------------------------------------------------------------
# Ratings override (Track A fallback source)
# ---------------------------------------------------------------------------

def load_ratings(directory: Optional[str] = None) -> dict:
    """{race_id: rating} overriding data/reference/races.json ratings."""
    directory = directory or config.MANUAL_DIR
    path = _find_file(FILES["ratings"], directory)
    if path is None:
        return {}
    fname = os.path.basename(path)
    from .schema import RATINGS
    out = {}
    for i, row in enumerate(_read_rows(path), start=2):
        rid = _check_race_id(row.get("race_id"), i, fname)
        rating = str(row.get("rating") or "").strip()
        if rating == "Toss-up":
            rating = "Tossup"
        if rating not in RATINGS:
            raise ManualDataError(
                "%s row %d: rating=%r must be one of %s"
                % (fname, i, row.get("rating"), "/".join(RATINGS)))
        out[rid] = rating
    return out
