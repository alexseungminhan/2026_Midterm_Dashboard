"""import_polls.py — convert transcribed NYT poll workbooks into data/manual/polls.

    python -m election2026 import-polls "Election Polls/Senate_Polls_2026.xlsx" \
        "Election Polls/Gubernatorial_Polls_2026.xlsx" 2026_House_poll.xlsx

Four workbook shapes are understood and may be imported together; the output
sheet is the union, so every monitored race's polls live in one file.

  TRACKER   `전체 여론조사` — the 2026-08-11 re-transcription in
            `Election Polls/`, one workbook per chamber, one row per trial
            heat, and — this is what the older sheets lack — a SAMPLE SIZE on
            every row. Covers the ten Senate and ten governor races the board
            actually shows. See "TRACKER WORKBOOKS" below.
  SENATE    `여론조사_전체` — one row per matchup, with the sheet's own
            `매치업 상태` (확정 / 가상 / 무효) and `라운드` columns.
  HOUSE     `전체 여론조사(통합)` — one row per matchup across the 18 tossup
            districts, plus one sheet per district whose header line records
            whether that district's primary has been held.
  GOVERNOR  `All polls` — one row per matchup across the five battleground
            states, with English headers and both an "as shown" NYT margin
            and the transcribed Dem %/Rep % to check it against.

The TRACKER workbooks SUPERSEDE the SENATE and GOVERNOR ones for the races
they cover. Do not pass both — `run()` writes the union, so the same poll
transcribed in two workbooks would be counted twice.

MATCHUP CONFIDENCE, not exclusion
---------------------------------
The old importer kept only 확정 rows. That was too strict: it left Michigan,
New Hampshire and Maine with no polls at all, so their consensus came from
betting markets alone even though the sheet held dozens of usable reads of
those states' partisan environment. Every usable row is now imported with a
matchup-confidence weight (manual.MATCHUP_WEIGHTS):

  confirmed      1.00  both nominees settled and polled by name
  generic_ballot 0.60  "Dem." vs "Rep." — no candidate quality, but it prices
                       exactly the D-vs-R event p_consensus is defined on
  hypothetical   0.35  a named matchup the primary has not settled
  withdrawn      0.15  a named matchup where a polled candidate has withdrawn

Still excluded, and these are exclusions of KIND rather than of confidence:

  * Intra-party primary polls (민주당 예비선거) and non-partisan top-two
    primary polls (예비선거(무당파)). "Rutinel +13 over Bird" is a contest
    between two Democrats; it carries no information about whether a Democrat
    beats a Republican, so there is no weight small enough to make it belong
    in p_consensus. Down-weighting a quantity that measures the wrong event
    does not make it the right event — it just makes the error smaller.
  * Nebraska (unless --include-independents). Dan Osborn runs as an
    independent, so Polymarket's "will a Democrat win Nebraska" market prices
    at 0.01 while the polls (Osborn vs Ricketts) imply 0.795. Blending them
    yields ~0.215, a number neither channel believes. Same principle as
    above: the two channels must price the same event.
  * Rows with no margin, no date, or an unresolvable candidate party.

TRIAL HEATS — one poll, one row
------------------------------
A pollster that asks five head-to-heads of one 800-person sample has run ONE
poll, not five. The sheets store a row per head-to-head, so emitting them
verbatim multiplies that pollster's weight by five inside aggregate_polls:
Marquette's July Wisconsin poll alone would outweigh every other 2026
Wisconsin poll combined. collapse_trial_heats() therefore folds the rows
sharing (race, pollster, end date) into one, keeping only the heats at the
STRONGEST matchup class the poll contains and averaging those margins. A poll
that tested the actual nominees alongside four also-rans is a read of the real
matchup; the also-rans add nothing to it. Where every heat is equally
speculative — as in Wisconsin, whose primary is not until 2026-08-11 — the
average across the tested Democrats IS the estimate of the partisan
environment, which is the quantity a hypothetical matchup can support.

This also applies to the House sheet, where twelve of its thirty-two general
election polls carry more than one heat.

RCV rounds — Alaska and Maine store "1st round" and "Final round" as separate
rows for one poll. The final round decides the seat, so it wins.

sample_size is not published on the NYT LIST screen, so the three older
workbooks leave it empty and manual.aggregate_polls falls back to a neutral
n=500 weight. The tracker workbooks carry it (read off each poll's own
screen), so their rows are weighted by sqrt(n) for real.

TRACKER WORKBOOKS — how a bare surname becomes a D-vs-R margin
--------------------------------------------------------------
The tracker sheets name candidates by surname only ("Peltola", "Sullivan"),
so three things have to be recovered before a row is a poll of the event
p_consensus is defined on:

1. WHICH PARTY. Each race gets one or two seed surnames (TRACKER_PARTY_SEEDS,
   taken from Polymarket's own party-labelled legs where it has them), and the
   rest are coloured by the same opposition-graph walk the House sheet uses:
   the two names in a head-to-head are on opposite sides. A race with no seed
   is REPORTED and dropped, never guessed — Montana, whose four surnames the
   market does not label, is the live example.
2. IS IT THE REAL MATCHUP. TRACKER_NOMINEES holds the pairs the market names
   as both nominees; a heat between those two is `confirmed`, a heat naming
   someone else is `withdrawn` (they lost the primary, which puts them off the
   ballot exactly as withdrawing would). Races the market prices with generic
   party legs get no entry and stay `hypothetical`.
3. IS IT EVEN A D-VS-R HEAT. California's May poll tested Becerra vs Porter
   and Hilton vs Bianco — intra-party heats, dropped for the same reason the
   House sheet drops primary polls. Same-party pairs are refused, not signed.

The margin is computed from the two transcribed shares and CHECKED against
the sheet's own margin text; a disagreement beyond GOVERNOR_MARGIN_TOLERANCE
refuses the row as a transcription error rather than averaging it in.
"""

from __future__ import annotations

import os
import re
from datetime import date, datetime
from typing import Optional

from . import config, manual

SENATE_SHEET = "여론조사_전체"
HOUSE_SHEET = "전체 여론조사(통합)"
GOVERNOR_SHEET = "All polls"

SENATE_COLUMNS = {
    "race_id": "race_id",
    "pollster": "여론조사 기관(Pollster)",
    "sponsor": "스폰서(Sponsor)",
    "lean": "스폰서 성향",
    "end_date": "종료일",
    "dem": "민주/도전 후보",
    "rep": "공화 후보",
    "margin": "마진(D기준, 계산)",
    "status": "매치업 상태",
    "round": "라운드",
}

# 매치업 상태 -> matchup-confidence class.
SENATE_STATUS_MATCHUP = {
    "확정": "confirmed",
    "가상": "hypothetical",
    "무효": "withdrawn",
}

FINAL_ROUND = "Final round"

# Races whose leading non-Republican is an independent, so the poll margin is
# not the "will a Democrat win" quantity p_consensus is defined on.
INDEPENDENT_RACES = {"sen-ne": "Dan Osborn runs as an independent"}

# House 조사 종류 values that do not price the D-vs-R event at all.
HOUSE_NON_GENERAL = {
    "민주당 예비선거": "an intra-party Democratic primary contest",
    "공화당 예비선거": "an intra-party Republican primary contest",
    "예비선거(무당파)": "a non-partisan top-two primary with more than two "
                        "candidates on one ballot",
}
HOUSE_GENERAL = "본선거"

# -- Tracker workbooks (Election Polls/, transcribed 2026-08-11) --------------
TRACKER_SHEET = "전체 여론조사"

# Both chambers' sheets carry the same information under different headers.
# The reader picks whichever map the header row satisfies.
TRACKER_SENATE_COLUMNS = {
    "state": "주", "pollster": "조사기관", "sponsor": "의뢰/후원",
    "field": "조사기간(원문)", "end_date": "조사 종료일",
    "sample_size": "표본수", "round": "라운드", "margin": "격차",
    "note": "비고",
    "c1": "후보1", "c1_pct": "후보1 %", "c2": "후보2", "c2_pct": "후보2 %",
}
TRACKER_GOVERNOR_COLUMNS = {
    "state": "주(State)", "pollster": "여론조사기관(Pollster)",
    "sponsor": "후원/의뢰(Sponsor)", "field": "조사기간(원문)",
    "end_date": "조사 종료일", "sample_size": "조사인원(표본수)",
    "round": "라운드", "margin": "격차(Margin)", "note": "비고",
    "c1": "1위 후보", "c1_pct": "1위 지지율(%)",
    "c2": "2위 후보", "c2_pct": "2위 지지율(%)",
}
TRACKER_SHAPES = (("senate", TRACKER_SENATE_COLUMNS),
                  ("governor", TRACKER_GOVERNOR_COLUMNS))

# One or two surnames per race, enough to colour that race's opposition graph.
# EVERY entry marked (market) is read off Polymarket's own party-labelled legs
# — "Jon Ossoff (D)", "Ashley Hinson (R)" — so the seed is data, not memory.
# The rest are sitting officeholders whose party is not in dispute, and each
# says which office is doing the vouching.
#
# A race that is not here is not guessed at: its rows are dropped with the
# surnames named, so the fix is to add a seed rather than to debug a silent
# sign error. sen-mt is exactly that case today — Polymarket prices Montana
# with bare "Republican"/"Independent"/"Democrat" legs, so nothing in the data
# says which of Alme/Bodnar/Bankhead/Austin is which.
TRACKER_PARTY_SEEDS = {
    "sen-ak": {"Peltola": "D", "Sullivan": "R"},        # (market)
    "sen-ga": {"Ossoff": "D"},                          # (market)
    "sen-ia": {"Hinson": "R", "Turek": "D"},            # (market)
    "sen-me": {"Jackson": "D", "Collins": "R"},         # (market)
    "sen-mi": {"El-Sayed": "D", "Rogers": "R"},         # (market)
    "sen-oh": {"Brown": "D", "Husted": "R"},            # (market)
    "sen-tx": {"Talarico": "D", "Paxton": "R"},         # (market)
    # Annie Andrews is the Democrat; every South Carolina row tests her
    # against a different Republican, which is what the graph needs.
    "sen-sc": {"Andrews": "D"},
    "gov-ak": {"Begich": "D", "Wilson": "R"},           # (market)
    # California is seeded EXHAUSTIVELY, and has to be. It is a top-two
    # primary state, so its pollsters test same-party heats (Becerra vs
    # Porter, Hilton vs Bianco) in the same poll as the D-vs-R ones — and the
    # opposition-graph walk assumes every heat crosses party lines. Colouring
    # California by propagation makes Porter a Republican and then reads
    # "Becerra +20 over Porter" as a 20-point Democratic lead. With all five
    # polled Californians seeded (market), those heats are recognized as
    # intra-party and dropped instead.
    "gov-ca": {"Becerra": "D", "Porter": "D", "Steyer": "D",
               "Hilton": "R", "Bianco": "R"},           # (market)
    "gov-ga": {"Bottoms": "D", "Jackson": "R"},         # (market)
    "gov-oh": {"Acton": "D", "Ramaswamy": "R"},         # (market)
    "gov-az": {"Hobbs": "D"},                           # sitting governor
    "gov-mi": {"Benson": "D"},                          # secretary of state
    "gov-mn": {"Klobuchar": "D"},                       # US senator
    "gov-ny": {"Hochul": "D"},                          # sitting governor
    "gov-ri": {"McKee": "D"},                           # sitting governor
    "gov-wi": {"Tiffany": "R"},                         # US representative
}

# Both nominees settled, as named by Polymarket's legs. A heat between these
# two is `confirmed`; a heat naming anyone else is `withdrawn`, because the
# primary is over and they are not on the November ballot.
#
# A race whose market still trades on generic party legs is deliberately
# absent — gov-wi, gov-ri, gov-mn, gov-ny, gov-az, gov-mi and sen-sc are not
# settled as far as the priced data goes, so their heats stay `hypothetical`.
TRACKER_NOMINEES = {
    "sen-ga": {"D": "Ossoff", "R": "Collins"},
    "sen-ia": {"D": "Turek", "R": "Hinson"},
    "sen-me": {"D": "Jackson", "R": "Collins"},
    "sen-mi": {"D": "El-Sayed", "R": "Rogers"},
    "sen-oh": {"D": "Brown", "R": "Husted"},
    "sen-tx": {"D": "Talarico", "R": "Paxton"},
    "gov-ca": {"D": "Becerra", "R": "Hilton"},
    "gov-ga": {"D": "Bottoms", "R": "Jackson"},
    "gov-oh": {"D": "Acton", "R": "Ramaswamy"},
    # Alaska has no nominees to settle: the top-four primary sends four
    # candidates to a ranked-choice general, so Peltola and Sullivan are both
    # certain to be on the November ballot no matter what August does. The
    # market agrees — it prices the two of them at 53/48 and every other
    # Alaskan at 0.1%.
    "sen-ak": {"D": "Peltola", "R": "Sullivan"},
}

# Rows whose leader is neither D nor R. gov-ak polls Bill Walker, an
# independent former governor, alongside the party candidates; a Walker heat
# is not the D-vs-R event, exactly as Nebraska's Osborn heats are not.
TRACKER_NON_MAJOR = {"Walker": "Bill Walker runs as an independent"}

FINAL_ROUND_LABELS = {"final round", "최종", "최종 라운드"}

# -- Governor workbook -------------------------------------------------------
GOVERNOR_COLUMNS = {
    "state": "State",
    "pollster": "Pollster",
    "sponsor": "Sponsor",
    "lean": "Sponsor lean",
    "field": "Field dates",
    "end_date": "End date",
    "dem": "Dem candidate",
    "dem_pct": "Dem %",
    "rep": "Rep candidate",
    "rep_pct": "Rep %",
    "margin": "Margin (as shown)",
    "leader": "Leader",
    "party": "Party",
}
# The sheet's "Dem - Rep" column is Dem % - Rep % restated, so it is not read:
# _governor_margin checks NYT's published margin against the raw toplines
# instead, which catches a mistyped share that the derived column would not.

# Date each state's 2026 gubernatorial NOMINATION was settled — the runoff
# where one was needed, not the first-round primary (Georgia's Republican
# nomination went to a 2026-06-16 runoff). Before this date every named
# matchup in that state is a hypothetical; after it, who the nominees are is
# read off the sheet's own post-primary polls rather than hard-coded here, so
# a candidate substitution needs no code change.
GOVERNOR_PRIMARY_DATES = {
    "GA": date(2026, 6, 16),
    "IA": date(2026, 6, 2),
    "NV": date(2026, 6, 9),
    "OH": date(2026, 5, 5),
    "WI": date(2026, 8, 11),
}

# The sheet carries NYT's own margin (computed on unrounded shares) alongside
# the rounded toplines, so the two legitimately differ by up to a point. More
# than that is a transcription error — a dropped digit or a mispasted row —
# and the row is refused rather than silently averaged in.
GOVERNOR_MARGIN_TOLERANCE = 1.5

# Generic party labels used in place of a candidate name. The Korean forms are
# how the tracker sheets spell the same thing ("일반 민주당 후보"); Rhode
# Island's sheet uses the bare English "Republican" for an unsettled nominee.
_GENERIC_D = {"dem.", "dem", "democrat", "democratic", "generic dem.",
              "democrat(일반)", "일반 민주당 후보", "민주당(일반)"}
_GENERIC_R = {"rep.", "rep", "republican", "gop", "generic rep.",
              "republican(일반)", "일반 공화당 후보", "공화당(일반)"}

_MONTHS = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
           "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


class PollImportError(ValueError):
    """The workbook could not be read as a poll sheet."""


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _sheet_rows(path: str, sheet: str) -> list:
    """[{header: value}] for one sheet, or [] when the sheet is absent."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    # The House sheets carry a title line and a blank line before the header;
    # the Senate sheet starts on the header. Find the header by looking for
    # the first row that names a pollster column.
    headers = None
    out = []
    for values in rows:
        if values is None or all(v is None for v in values):
            continue
        cells = [str(v).strip() if v is not None else "" for v in values]
        if headers is None:
            if any(c.startswith("여론조사 기관") or c == "race_id"
                   for c in cells):
                headers = cells
            continue
        out.append({h: v for h, v in zip(headers, values) if h})
    if headers is None:
        raise PollImportError("%s sheet %r has no recognizable header row"
                              % (os.path.basename(path), sheet))
    return out


def _norm(value) -> str:
    return str(value or "").strip()


# Generational suffixes are not surnames; without stripping them "Thomas Kean
# Jr." matches on "Jr." and the incumbent seeds no party at all.
_NAME_SUFFIXES = {"jr", "sr", "ii", "iii", "iv", "v"}


def _name_tokens(name: str) -> list:
    return [t for t in re.split(r"[\s.,]+", _norm(name).lower())
            if t and t not in _NAME_SUFFIXES]


def _party_of_label(name: str) -> Optional[str]:
    # The governor sheet spells its generic ballot "Democrat (generic)"; the
    # House sheet spells the same thing "Dem.". Strip the parenthetical so one
    # table covers both.
    token = re.sub(r"\s*\([^)]*\)\s*", " ", _norm(name)).strip().lower()
    if token in _GENERIC_D:
        return "D"
    if token in _GENERIC_R:
        return "R"
    return None


def _is_generic(name: str) -> bool:
    return _party_of_label(name) is not None


def parse_field_dates(text: str, default_year: int = 2026) -> Optional[date]:
    """End date of an NYT field-period string.

    Handles "June 8-11", "Nov. 15-17, 2025", "June 27 - July 1", "May 5-6".
    The last month named and the last day number win; a trailing four-digit
    year overrides the default (the sheet omits it for the current cycle).
    """
    text = _norm(text)
    if not text:
        return None
    year_match = re.search(r"\b(20\d{2})\b", text)
    year = int(year_match.group(1)) if year_match else default_year
    body = text[:year_match.start()] if year_match else text
    months = [_MONTHS[m.group(0)[:3].lower()]
              for m in re.finditer(r"[A-Za-z]{3,9}\.?", body)
              if m.group(0)[:3].lower() in _MONTHS]
    days = [int(d) for d in re.findall(r"\b(\d{1,2})\b", body)]
    if not months or not days:
        return None
    try:
        return date(year, months[-1], days[-1])
    except ValueError:
        return None


def parse_margin(text: str) -> Optional[tuple]:
    """('Even'|leader token, points) from an NYT margin string.

    "Mendoza +2" -> ("Mendoza", 2.0); "Democrat +12" -> ("Democrat", 12.0);
    "Even" -> (None, 0.0).
    """
    text = _norm(text)
    if not text:
        return None
    if text.lower() in ("even", "tie", "tied", "0"):
        return None, 0.0
    m = re.match(r"^(.*?)\s*\+\s*([\d.]+)$", text)
    if not m:
        return None
    who = m.group(1).strip()
    try:
        return (who or None), float(m.group(2))
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Trial heats -> one poll
# ---------------------------------------------------------------------------

# Strongest first, so min() over this picks the most informative class a poll
# contains. Derived from the WEIGHTS rather than from the declaration order of
# MATCHUP_KINDS, so re-tuning a weight cannot silently reorder this.
_MATCHUP_RANK = {kind: -weight
                 for kind, weight in manual.MATCHUP_WEIGHTS.items()}


def collapse_trial_heats(polls: list) -> tuple:
    """Fold a poll's several head-to-heads into one row. -> (rows, n_absorbed).

    n_absorbed counts every input row that did not survive as its own output
    row — whether it was outranked by a stronger matchup or averaged into one
    at the same confidence — so the caller's "N polls of M rows" arithmetic
    stays exact. Which of the two happened is recorded per row in `notes`.

    Rows must carry the private keys `_note_pre` (shared note fragments),
    `_heat` (this head-to-head's one-line description) and `_source`; all
    three are consumed here and never reach the output sheet. See the module
    docstring for why this is not simply "keep them all".
    """
    groups: dict = {}
    order: list = []
    for poll in polls:
        key = (poll["race_id"], poll["pollster"] or "", poll["date"])
        if key not in groups:
            order.append(key)
        groups.setdefault(key, []).append(poll)

    out, absorbed = [], 0
    for key in order:
        heats = groups[key]
        best = min(_MATCHUP_RANK[h["matchup"]] for h in heats)
        kept = [h for h in heats if _MATCHUP_RANK[h["matchup"]] == best]
        absorbed += len(heats) - 1
        head = dict(kept[0])
        head["margin_dem"] = round(
            sum(h["margin_dem"] for h in kept) / len(kept), 2)

        note = list(head.pop("_note_pre"))
        note.append(" | ".join(h["_heat"] for h in kept))
        if len(heats) > 1:
            note.append("%d trial heats in one poll -> 1 row (%s)"
                        % (len(heats),
                           "averaged the %d at %s" % (len(kept), head["matchup"])
                           if len(kept) > 1
                           else "kept the only one at %s" % head["matchup"]))
        if head["matchup"] != "confirmed":
            note.append("%s @ w=%.2f"
                        % (head["matchup"], head["weight"]))
        note.append(head.pop("_source"))
        head.pop("_heat")
        head["notes"] = "; ".join(n for n in note if n)
        out.append(head)
    return out, absorbed


# ---------------------------------------------------------------------------
# Senate workbook
# ---------------------------------------------------------------------------

def _senate_race_id(raw: str) -> Optional[str]:
    """senate_GA_2026 -> sen-ga, if that race is monitored."""
    parts = _norm(raw).split("_")
    if len(parts) != 3 or parts[0] != "senate":
        return None
    rid = "sen-%s" % parts[1].lower()
    return rid if rid in config.RACES else None


def read_senate_rows(path: str) -> list:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if SENATE_SHEET not in wb.sheetnames:
        raise PollImportError(
            "%s has no %r sheet (found: %s)"
            % (os.path.basename(path), SENATE_SHEET, ", ".join(wb.sheetnames)))
    ws = wb[SENATE_SHEET]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    missing = [c for c in SENATE_COLUMNS.values() if c not in headers]
    if missing:
        raise PollImportError("%s is missing column(s): %s"
                              % (os.path.basename(path), ", ".join(missing)))
    index = {key: headers.index(col) for key, col in SENATE_COLUMNS.items()}
    out = []
    for values in rows:
        if values is None or all(v is None for v in values):
            continue
        out.append({key: values[i] if i < len(values) else None
                    for key, i in index.items()})
    return out


def convert_senate(rows: list, include_independents: bool = False) -> tuple:
    """(poll rows, {reason: count} of what was dropped)."""
    dropped: dict = {}

    def drop(reason):
        dropped[reason] = dropped.get(reason, 0) + 1

    best = {}
    for row in rows:
        rid = _senate_race_id(row["race_id"])
        if rid is None:
            drop("not a monitored senate race (incl. the sheet's footer rows)")
            continue
        if rid in INDEPENDENT_RACES and not include_independents:
            drop("%s: %s, so the margin is not a D-vs-R probability and the "
                 "betting market prices a different event"
                 % (rid, INDEPENDENT_RACES[rid]))
            continue
        matchup = SENATE_STATUS_MATCHUP.get(_norm(row["status"]))
        if matchup is None:
            drop("매치업 상태 is %s — not one of %s"
                 % (_norm(row["status"]) or "blank",
                    "/".join(SENATE_STATUS_MATCHUP)))
            continue
        if row["margin"] is None or not row["end_date"]:
            drop("no margin or no end date")
            continue
        # A hypothetical matchup between two GENERIC party labels is the
        # named-party generic ballot, which needs no nominee and therefore
        # deserves the generic_ballot weight rather than the hypothetical one.
        if matchup == "hypothetical" and _is_generic(row["dem"]) \
                and _is_generic(row["rep"]):
            matchup = "generic_ballot"

        # One poll = one (race, pollster, end date). RCV rounds collapse here.
        key = (rid, _norm(row["pollster"]), str(row["end_date"])[:10])
        if key in best:
            if best[key]["round"] == FINAL_ROUND:
                drop("duplicate RCV round (kept the final round)")
                continue
            drop("duplicate RCV round (kept the final round)")
        best[key] = dict(row, _matchup=matchup)

    polls = []
    for (rid, pollster, when), row in sorted(best.items()):
        note = []
        if row["round"]:
            note.append(str(row["round"]))
        if row["lean"]:
            note.append("%s-sponsored" % row["lean"])
        elif row["sponsor"]:
            note.append("sponsor: %s" % row["sponsor"])
        if row["_matchup"] != "confirmed":
            note.append("매치업 %s -> %s @ w=%.2f"
                        % (_norm(row["status"]), row["_matchup"],
                           manual.MATCHUP_WEIGHTS[row["_matchup"]]))
        if rid in INDEPENDENT_RACES:
            note.append("%s — margin_dem is the independent minus the "
                        "Republican, NOT a D-vs-R margin"
                        % INDEPENDENT_RACES[rid])
        note.append("NYT via 2026_senate_poll.xlsx; sample size not published")
        polls.append({
            "race_id": rid,
            "pollster": pollster or None,
            "date": when,
            "sample_size": None,
            "margin_dem": float(row["margin"]),
            "matchup": row["_matchup"],
            "weight": manual.MATCHUP_WEIGHTS[row["_matchup"]],
            "notes": "; ".join(note),
        })
    return polls, dropped


# ---------------------------------------------------------------------------
# House workbook
# ---------------------------------------------------------------------------

def _house_label_map() -> dict:
    """{'Ariz. 1': 'ho-az1'} over the monitored House board."""
    return {m["label"]: rid for rid, m in config.RACES.items()
            if m["chamber"] == "house"}


def read_house_primary_status(path: str) -> dict:
    """{district label: True if its primary has already been held}.

    Read from each district sheet's second line, which the transcription
    records as e.g. "예비선거 5월 19일 실시" (held) or "8월 4일 예정"
    (scheduled). A district whose sheet says nothing is treated as NOT held,
    which is the conservative reading: an unsettled nominee down-weights.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        if name in (HOUSE_SHEET,) or name.startswith("전체"):
            continue
        ws = wb[name]
        header = ""
        for i, values in enumerate(ws.iter_rows(values_only=True)):
            if i > 2:
                break
            if values and values[0]:
                header += " " + str(values[0])
        out[name] = "실시" in header and "예정" not in header
    return out


def _resolve_parties(rows: list, meta: dict) -> dict:
    """{candidate name: 'D'|'R'} for one district's general-election rows.

    Seeded from the generic party labels ("Dem."/"Rep.") and the incumbent,
    then propagated across the opposition graph: the two names in a
    head-to-head matchup are on opposite sides, so colouring one colours the
    other. Names that stay uncoloured are reported, not guessed.
    """
    parties: dict = {}
    edges = []
    for row in rows:
        a, b = _norm(row["c1"]), _norm(row["c2"])
        for name in (a, b):
            p = _party_of_label(name)
            if p:
                parties[name] = p
        if a and b:
            edges.append((a, b))

    incumbent = _norm(meta.get("incumbent"))
    if incumbent and meta.get("incumbent_party"):
        # Polls list a bare surname ("Kean"), races.json a full name with a
        # generational suffix ("Thomas Kean Jr."). Match on the last
        # NON-SUFFIX token, or the suffix itself becomes the surname and the
        # incumbent seeds nothing.
        tokens = _name_tokens(incumbent)
        last = tokens[-1] if tokens else None
        for name in {n for e in edges for n in e}:
            if last and last in _name_tokens(name):
                parties[name] = meta["incumbent_party"]

    flip = {"D": "R", "R": "D"}
    for _ in range(len(edges) + 1):          # propagate to a fixpoint
        changed = False
        for a, b in edges:
            if a in parties and b not in parties:
                parties[b] = flip[parties[a]]
                changed = True
            elif b in parties and a not in parties:
                parties[a] = flip[parties[b]]
                changed = True
        if not changed:
            break
    return parties


def read_house_rows(path: str) -> list:
    rows = _sheet_rows(path, HOUSE_SHEET)
    if not rows:
        raise PollImportError(
            "%s has no %r sheet" % (os.path.basename(path), HOUSE_SHEET))
    out = []
    for row in rows:
        out.append({
            "label": _norm(row.get("지역구")),
            "kind": _norm(row.get("조사 종류")),
            "pollster": _norm(row.get("여론조사 기관")),
            "field": _norm(row.get("조사 기간")),
            "sponsor": _norm(row.get("스폰서")),
            "lean": _norm(row.get("스폰서 성향")),
            "c1": _norm(row.get("후보 1")),
            "c2": _norm(row.get("후보 2")),
            "margin": _norm(row.get("마진")),
        })
    return [r for r in out if r["label"]]


def convert_house(rows: list, primary_held: Optional[dict] = None) -> tuple:
    """(poll rows, {reason: count} of what was dropped)."""
    primary_held = primary_held or {}
    labels = _house_label_map()
    dropped: dict = {}

    def drop(reason):
        dropped[reason] = dropped.get(reason, 0) + 1

    by_label: dict = {}
    for row in rows:
        if row["label"] not in labels:
            drop("not one of the 18 monitored districts (%s)" % row["label"])
            continue
        if row["kind"] in HOUSE_NON_GENERAL:
            drop("%s — %s, which does not price the D-vs-R event p_consensus "
                 "is defined on" % (row["kind"], HOUSE_NON_GENERAL[row["kind"]]))
            continue
        if row["kind"] != HOUSE_GENERAL:
            drop("조사 종류 %r is not %s" % (row["kind"], HOUSE_GENERAL))
            continue
        by_label.setdefault(row["label"], []).append(row)

    polls = []
    for label, district_rows in sorted(by_label.items()):
        rid = labels[label]
        meta = config.RACES[rid]
        parties = _resolve_parties(district_rows, meta)

        # A district counts as settled only when its primary has actually been
        # held AND the sheet polls exactly one name per side. Two names on a
        # side means the sheet itself is unsure who the nominee is.
        named = {"D": set(), "R": set()}
        for row in district_rows:
            for name in (row["c1"], row["c2"]):
                p = parties.get(name)
                if p and not _is_generic(name):
                    named[p].add(name)
        settled = (primary_held.get(label, False)
                   and len(named["D"]) <= 1 and len(named["R"]) <= 1)

        for row in district_rows:
            when = parse_field_dates(row["field"])
            if when is None:
                drop("unparseable 조사 기간 %r" % row["field"])
                continue
            parsed = parse_margin(row["margin"])
            if parsed is None:
                drop("unparseable 마진 %r" % row["margin"])
                continue
            leader, points = parsed
            if points == 0:
                margin_dem = 0.0
            else:
                party = (_party_of_label(leader) if leader else None) \
                    or parties.get(leader)
                if party is None:
                    drop("%s: cannot tell which party %r belongs to"
                         % (label, leader))
                    continue
                margin_dem = points if party == "D" else -points

            generic = _is_generic(row["c1"]) and _is_generic(row["c2"])
            if generic:
                matchup = "generic_ballot"
            elif settled:
                matchup = "confirmed"
            else:
                matchup = "hypothetical"

            note_pre = []
            if row["lean"]:
                note_pre.append("%s-sponsored" % row["lean"])
            elif row["sponsor"]:
                note_pre.append("sponsor: %s" % row["sponsor"])
            polls.append({
                "race_id": rid,
                "pollster": row["pollster"] or None,
                "date": when.isoformat(),
                "sample_size": None,
                "margin_dem": margin_dem,
                "matchup": matchup,
                "weight": manual.MATCHUP_WEIGHTS[matchup],
                "_note_pre": note_pre,
                "_heat": "%s vs %s D%+.1f" % (row["c1"], row["c2"], margin_dem),
                "_source": "NYT via 2026_House_poll.xlsx; sample size not "
                           "published",
            })

    polls, absorbed = collapse_trial_heats(polls)
    if absorbed:
        dropped["extra trial heats folded into their poll's row — outranked "
                "by a stronger matchup, or averaged in at the same "
                "confidence (one poll = one row)"] = absorbed
    return polls, dropped


# ---------------------------------------------------------------------------
# Governor workbook
# ---------------------------------------------------------------------------

_STATE_CODES = {name.lower(): code for code, name in config.STATE_NAMES.items()}


def _governor_race_id(state_name: str) -> Optional[str]:
    """'Wisconsin' -> gov-wi, if that race is monitored."""
    code = _STATE_CODES.get(_norm(state_name).lower())
    if code is None:
        return None
    rid = "gov-%s" % code.lower()
    return rid if rid in config.RACES else None


def read_governor_rows(path: str) -> list:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if GOVERNOR_SHEET not in wb.sheetnames:
        raise PollImportError(
            "%s has no %r sheet (found: %s)"
            % (os.path.basename(path), GOVERNOR_SHEET,
               ", ".join(wb.sheetnames)))
    ws = wb[GOVERNOR_SHEET]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    missing = [c for c in GOVERNOR_COLUMNS.values() if c not in headers]
    if missing:
        raise PollImportError("%s sheet %r is missing column(s): %s"
                              % (os.path.basename(path), GOVERNOR_SHEET,
                                 ", ".join(missing)))
    index = {key: headers.index(col) for key, col in GOVERNOR_COLUMNS.items()}
    out = []
    for values in rows:
        if values is None or all(v is None for v in values):
            continue
        out.append({key: values[i] if i < len(values) else None
                    for key, i in index.items()})
    return out


def _governor_end_date(row: dict) -> Optional[date]:
    """The sheet's own End date column, falling back to the field-date text."""
    value = row.get("end_date")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _norm(value)
    if text:
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            pass
    return parse_field_dates(row.get("field"))


def _governor_nominees(rows: list, primary: Optional[date],
                       as_of: date) -> tuple:
    """({'D': name|None, 'R': name|None}, primary_held) for one state.

    Read off the sheet rather than hard-coded: once the nomination is settled,
    the candidates a pollster keeps testing ARE the nominees. A side whose
    post-primary polls name more than one candidate equally often is left
    unresolved — the sheet is not telling us who won, and guessing would
    promote an also-ran's numbers to `confirmed`.
    """
    if primary is None or primary > as_of:
        return {"D": None, "R": None}, False
    counts: dict = {"D": {}, "R": {}}
    for row in rows:
        when = _governor_end_date(row)
        if when is None or when < primary:
            continue
        for side, name in (("D", _norm(row["dem"])), ("R", _norm(row["rep"]))):
            if name and not _is_generic(name):
                counts[side][name] = counts[side].get(name, 0) + 1
    nominees = {}
    for side, tally in counts.items():
        ranked = sorted(tally.items(), key=lambda kv: (-kv[1], kv[0]))
        nominees[side] = (ranked[0][0]
                          if ranked and (len(ranked) == 1
                                         or ranked[0][1] > ranked[1][1])
                          else None)
    return nominees, True


def _governor_margin(row: dict) -> tuple:
    """(margin_dem, reason_if_refused).

    Prefers NYT's published margin, whose sign comes from the sheet's Party
    column, and checks it against the transcribed Dem % - Rep %.
    """
    shown = _norm(row["margin"])
    party = _norm(row["party"]).upper()
    parsed = parse_margin(shown)

    from_pct = None
    if row["dem_pct"] is not None and row["rep_pct"] is not None:
        try:
            from_pct = (float(row["dem_pct"]) - float(row["rep_pct"])) * 100.0
        except (TypeError, ValueError):
            from_pct = None

    if parsed is None:
        if from_pct is None:
            return None, "unparseable margin %r and no Dem %%/Rep %%" % shown
        return round(from_pct, 2), None

    leader, points = parsed
    if points == 0:
        margin = 0.0
    elif party in ("D", "R"):
        margin = points if party == "D" else -points
    else:
        # No usable Party cell: fall back to matching the leader's name
        # against the two candidate columns.
        if leader and _norm(leader) == _norm(row["dem"]):
            margin = points
        elif leader and _norm(leader) == _norm(row["rep"]):
            margin = -points
        else:
            return None, ("margin %r names a leader (%s) that is neither the "
                          "Dem nor the Rep column and Party=%r"
                          % (shown, leader, _norm(row["party"])))

    if from_pct is not None and abs(margin - from_pct) > GOVERNOR_MARGIN_TOLERANCE:
        return None, ("margin %r (%+.1f) disagrees with Dem %% - Rep %% "
                      "(%+.1f) by more than %.1f points — transcription error"
                      % (shown, margin, from_pct, GOVERNOR_MARGIN_TOLERANCE))
    return margin, None


def convert_governor(rows: list, as_of: Optional[date] = None) -> tuple:
    """(poll rows, {reason: count} of what was dropped)."""
    as_of = as_of or date.today()
    dropped: dict = {}

    def drop(reason):
        dropped[reason] = dropped.get(reason, 0) + 1

    by_race: dict = {}
    for row in rows:
        rid = _governor_race_id(row["state"])
        if rid is None:
            drop("not one of the %d monitored governor races (%s)"
                 % (sum(1 for m in config.RACES.values()
                        if m["chamber"] == "governor"), _norm(row["state"])))
            continue
        by_race.setdefault(rid, []).append(row)

    polls = []
    for rid, state_rows in sorted(by_race.items()):
        state = config.RACES[rid]["state"]
        primary = GOVERNOR_PRIMARY_DATES.get(state)
        nominees, held = _governor_nominees(state_rows, primary, as_of)

        for row in state_rows:
            when = _governor_end_date(row)
            if when is None:
                drop("%s: no usable end date (End date=%r, Field dates=%r)"
                     % (rid, row["end_date"], _norm(row["field"])))
                continue
            margin, refused = _governor_margin(row)
            if refused is not None:
                drop("%s: %s" % (rid, refused))
                continue

            dem, rep = _norm(row["dem"]), _norm(row["rep"])
            if _is_generic(dem) or _is_generic(rep):
                matchup = "generic_ballot"
            elif not held or nominees["D"] is None or nominees["R"] is None:
                matchup = "hypothetical"
            elif dem == nominees["D"] and rep == nominees["R"]:
                matchup = "confirmed"
            else:
                # The nomination is settled and this is not it: a candidate
                # who lost the primary is off the ballot exactly as a
                # withdrawn one is.
                matchup = "withdrawn"

            note_pre = []
            if row["lean"]:
                note_pre.append("%s-sponsored" % _norm(row["lean"]))
            elif row["sponsor"]:
                note_pre.append("sponsor: %s" % _norm(row["sponsor"]))
            if matchup == "hypothetical" and primary and primary > as_of:
                note_pre.append("nomination unsettled until %s"
                                % primary.isoformat())
            polls.append({
                "race_id": rid,
                "pollster": _norm(row["pollster"]) or None,
                "date": when.isoformat(),
                "sample_size": None,
                "margin_dem": margin,
                "matchup": matchup,
                "weight": manual.MATCHUP_WEIGHTS[matchup],
                "_note_pre": note_pre,
                "_heat": "%s vs %s D%+.1f" % (dem, rep, margin),
                "_source": "NYT via Gubernatorial_Polls_2026.xlsx; sample "
                           "size not published",
            })

    polls, absorbed = collapse_trial_heats(polls)
    if absorbed:
        dropped["extra trial heats folded into their poll's row — outranked "
                "by a stronger matchup, or averaged in at the same "
                "confidence (one poll = one row)"] = absorbed
    return polls, dropped


# ---------------------------------------------------------------------------
# Tracker workbooks (Election Polls/)
# ---------------------------------------------------------------------------

def read_tracker_rows(path: str) -> tuple:
    """(chamber, [row dicts]) for a tracker workbook, or (None, []).

    (None, []) means the sheet is there but its headers are neither shape —
    the caller then reports the workbook rather than importing half of it.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if TRACKER_SHEET not in wb.sheetnames:
        return None, []
    ws = wb[TRACKER_SHEET]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return None, []

    for chamber, columns in TRACKER_SHAPES:
        if all(col in headers for col in columns.values()):
            index = {key: headers.index(col) for key, col in columns.items()}
            out = []
            for values in rows:
                if values is None or all(v is None for v in values):
                    continue
                row = {key: values[i] if i < len(values) else None
                       for key, i in index.items()}
                if _norm(row["state"]):
                    out.append(row)
            return chamber, out
    return None, []


def _tracker_state_code(raw: str) -> Optional[str]:
    """'Alaska (알래스카)' -> 'AK'. The Senate sheet annotates, the governor
    sheet does not; both are handled by cutting at the parenthesis."""
    name = re.sub(r"\s*\(.*$", "", _norm(raw)).strip()
    return _STATE_CODES.get(name.lower())


def _tracker_end_date(row: dict) -> Optional[date]:
    value = row.get("end_date")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _norm(value)
    if text:
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            pass
    return parse_field_dates(row.get("field"))


def _tracker_parties(rows: list, seeds: dict) -> dict:
    """{surname: 'D'|'R'} for one race, seeded then propagated.

    Same walk as _resolve_parties, but seeded from TRACKER_PARTY_SEEDS rather
    than from races.json — the tracker sheets cover races that reference file
    has never heard of (California governor among them).
    """
    parties = dict(seeds)
    edges = []
    for row in rows:
        a, b = _norm(row["c1"]), _norm(row["c2"])
        for name in (a, b):
            p = _party_of_label(name)
            if p:
                parties[name] = p
        if a and b:
            edges.append((a, b))

    flip = {"D": "R", "R": "D"}
    for _ in range(len(edges) + 1):
        changed = False
        for a, b in edges:
            if a in parties and b not in parties:
                parties[b] = flip[parties[a]]
                changed = True
            elif b in parties and a not in parties:
                parties[a] = flip[parties[b]]
                changed = True
        if not changed:
            break
    return parties


def _tracker_margin(row: dict, dem_pct, rep_pct) -> tuple:
    """(margin_dem, reason_if_refused) from the two transcribed shares.

    The shares are the primary source and the sheet's own margin text is the
    check on them — the reverse of the governor workbook, where the margin
    column was NYT's own and the shares were the transcription. Here both are
    transcribed off the same screen, and the shares carry the sign
    unambiguously once the parties are known.
    """
    if dem_pct is None or rep_pct is None:
        return None, "no share transcribed for one of the two candidates"
    try:
        margin = float(dem_pct) - float(rep_pct)
    except (TypeError, ValueError):
        return None, "unreadable shares %r / %r" % (dem_pct, rep_pct)

    shown = _norm(row["margin"])
    # "N/A" is how the sheet records a round the pollster did not publish.
    if shown.upper() in ("N/A", "NA", "-"):
        return None, "margin is N/A — the pollster did not publish this round"
    parsed = parse_margin(re.sub(r"동률\(Even\)|동률", "Even", shown))
    if parsed is not None:
        _, points = parsed
        if abs(abs(margin) - points) > GOVERNOR_MARGIN_TOLERANCE:
            return None, ("margin text %r (%.1f points) disagrees with the "
                          "transcribed shares (%+.1f) by more than %.1f — "
                          "transcription error"
                          % (shown, points, margin,
                             GOVERNOR_MARGIN_TOLERANCE))
    return round(margin, 2), None


def convert_tracker(chamber: str, rows: list,
                    include_independents: bool = False) -> tuple:
    """(poll rows, {reason: count} of what was dropped)."""
    dropped: dict = {}

    def drop(reason):
        dropped[reason] = dropped.get(reason, 0) + 1

    prefix = {"senate": "sen", "governor": "gov"}[chamber]
    by_race: dict = {}
    for row in rows:
        code = _tracker_state_code(row["state"])
        if code is None:
            drop("unrecognized state %r" % _norm(row["state"]))
            continue
        rid = "%s-%s" % (prefix, code.lower())
        if rid in INDEPENDENT_RACES and not include_independents:
            drop("%s: %s, so the margin is not a D-vs-R probability and the "
                 "betting market prices a different event"
                 % (rid, INDEPENDENT_RACES[rid]))
            continue
        by_race.setdefault(rid, []).append(row)

    polls = []
    for rid, race_rows in sorted(by_race.items()):
        seeds = TRACKER_PARTY_SEEDS.get(rid)
        if not seeds:
            names = sorted(({_norm(r["c1"]) for r in race_rows}
                            | {_norm(r["c2"]) for r in race_rows}) - {""})
            dropped["%s: no party seed, so %s cannot be signed D-vs-R — add "
                    "one to TRACKER_PARTY_SEEDS" % (rid, "/".join(names))] = \
                len(race_rows)
            continue
        parties = _tracker_parties(race_rows, seeds)
        nominees = TRACKER_NOMINEES.get(rid)

        # RCV: one poll's final round supersedes its own first round.
        finals = {(_norm(r["pollster"]), str(_tracker_end_date(r)))
                  for r in race_rows
                  if _norm(r["round"]).lower() in FINAL_ROUND_LABELS
                  and _tracker_margin(
                      r, r["c1_pct"], r["c2_pct"])[0] is not None}

        for row in race_rows:
            when = _tracker_end_date(row)
            if when is None:
                drop("%s: no usable end date (%r / %r)"
                     % (rid, row["end_date"], _norm(row["field"])))
                continue
            first, second = _norm(row["c1"]), _norm(row["c2"])
            if not first or not second:
                drop("%s: row names fewer than two candidates" % rid)
                continue
            is_final = _norm(row["round"]).lower() in FINAL_ROUND_LABELS
            if not is_final and (_norm(row["pollster"]), str(when)) in finals:
                drop("first round superseded by the same poll's final round "
                     "(ranked-choice)")
                continue

            blocked = [n for n in (first, second) if n in TRACKER_NON_MAJOR]
            if blocked:
                drop("%s: %s" % (rid, TRACKER_NON_MAJOR[blocked[0]]))
                continue
            p1, p2 = parties.get(first), parties.get(second)
            if p1 is None or p2 is None:
                unknown = first if p1 is None else second
                drop("%s: cannot tell which party %r belongs to" % (rid, unknown))
                continue
            if p1 == p2:
                drop("%s: %s vs %s is an intra-party heat, not the D-vs-R "
                     "event" % (rid, first, second))
                continue

            dem, rep = (first, second) if p1 == "D" else (second, first)
            dem_pct = row["c1_pct"] if p1 == "D" else row["c2_pct"]
            rep_pct = row["c2_pct"] if p1 == "D" else row["c1_pct"]
            margin, refused = _tracker_margin(row, dem_pct, rep_pct)
            if refused is not None:
                drop("%s: %s" % (rid, refused))
                continue

            if _is_generic(dem) or _is_generic(rep):
                matchup = "generic_ballot"
            elif nominees is None:
                matchup = "hypothetical"
            elif dem == nominees["D"] and rep == nominees["R"]:
                matchup = "confirmed"
            else:
                matchup = "withdrawn"

            n = row["sample_size"]
            try:
                n = int(n) if n not in (None, "") else None
            except (TypeError, ValueError):
                n = None

            note_pre = []
            if row["sponsor"]:
                note_pre.append("sponsor: %s" % _norm(row["sponsor"]))
            if is_final:
                note_pre.append("ranked-choice final round")
            if _norm(row["note"]):
                note_pre.append(_norm(row["note"]))
            polls.append({
                "race_id": rid,
                "pollster": _norm(row["pollster"]) or None,
                "date": when.isoformat(),
                "sample_size": n,
                "margin_dem": margin,
                "matchup": matchup,
                "weight": manual.MATCHUP_WEIGHTS[matchup],
                "_note_pre": note_pre,
                "_heat": "%s vs %s D%+.1f" % (dem, rep, margin),
                "_source": "NYT via Election Polls/ (2026-08-11)",
            })

    polls, absorbed = collapse_trial_heats(polls)
    if absorbed:
        dropped["extra trial heats folded into their poll's row — outranked "
                "by a stronger matchup, or averaged in at the same "
                "confidence (one poll = one row)"] = absorbed
    return polls, dropped


# ---------------------------------------------------------------------------
# Dispatch + write
# ---------------------------------------------------------------------------

def convert_workbook(path: str, include_independents: bool = False) -> tuple:
    """Detect the workbook shape and convert it."""
    from openpyxl import load_workbook

    sheets = load_workbook(path, read_only=True, data_only=True).sheetnames
    if TRACKER_SHEET in sheets:
        chamber, rows = read_tracker_rows(path)
        if chamber is None:
            raise PollImportError(
                "%s has a %r sheet but neither the Senate nor the governor "
                "header set — see TRACKER_SENATE_COLUMNS"
                % (os.path.basename(path), TRACKER_SHEET))
        return convert_tracker(chamber, rows, include_independents)
    if SENATE_SHEET in sheets:
        return convert_senate(read_senate_rows(path), include_independents)
    if HOUSE_SHEET in sheets:
        return convert_house(read_house_rows(path),
                             read_house_primary_status(path))
    if GOVERNOR_SHEET in sheets:
        return convert_governor(read_governor_rows(path))
    raise PollImportError(
        "%s has none of the %r / %r / %r sheets (found: %s)"
        % (os.path.basename(path), SENATE_SHEET, HOUSE_SHEET, GOVERNOR_SHEET,
           ", ".join(sheets)))


def write(polls: list, directory: Optional[str] = None) -> str:
    from openpyxl import Workbook

    directory = directory or config.MANUAL_DIR
    os.makedirs(directory, exist_ok=True)
    headers = manual.TEMPLATES["polls_template.xlsx"]
    wb = Workbook()
    ws = wb.active
    ws.title = "polls"
    ws.append(headers)
    for poll in polls:
        ws.append([poll.get(h) for h in headers])
    path = os.path.join(directory, "%s.xlsx" % manual.FILES["polls"])
    wb.save(path)
    return path


def run(sources, directory: Optional[str] = None,
        include_independents: bool = False) -> str:
    """Import one or more workbooks into a single data/manual/polls.xlsx.

    The output is the UNION across workbooks, so re-importing the Senate sheet
    alone would drop the House polls — always pass every source together.
    """
    if isinstance(sources, str):
        sources = [sources]

    polls: list = []
    dropped: dict = {}
    read_total = 0
    for source in sources:
        rows_dropped: dict
        converted, rows_dropped = convert_workbook(source, include_independents)
        base = os.path.basename(source)
        read_total += len(converted) + sum(rows_dropped.values())
        polls.extend(converted)
        for reason, n in rows_dropped.items():
            key = "%s: %s" % (base, reason)
            dropped[key] = dropped.get(key, 0) + n
        print("[import-polls] %s -> %d usable polls" % (base, len(converted)))

    if not polls:
        raise PollImportError(
            "no usable polls from %s — every row was filtered out"
            % ", ".join(os.path.basename(s) for s in sources))

    polls.sort(key=lambda p: (p["race_id"], p["date"], p["pollster"] or ""))
    path = write(polls, directory)

    by_race: dict = {}
    for poll in polls:
        by_race.setdefault(poll["race_id"], []).append(poll)

    for reason, n in sorted(dropped.items(), key=lambda kv: -kv[1]):
        print("[import-polls]   dropped %4d — %s" % (n, reason))
    print("[import-polls] wrote %d polls (of %d rows) -> %s"
          % (len(polls), read_total, path))

    by_matchup: dict = {}
    for poll in polls:
        by_matchup[poll["matchup"]] = by_matchup.get(poll["matchup"], 0) + 1
    print("[import-polls] matchup mix: %s"
          % ", ".join("%s=%d (w=%.2f)"
                      % (k, by_matchup[k], manual.MATCHUP_WEIGHTS[k])
                      for k in manual.MATCHUP_KINDS if k in by_matchup))

    for rid in sorted(by_race):
        rows = by_race[rid]
        margin = manual.aggregate_polls(rows)
        print("[import-polls]   %-8s %2d polls, aggregate margin D%+.1f, "
              "matchup confidence %.2f"
              % (rid, len(rows), margin, manual.poll_confidence(rows)))

    silent = sorted(r for r in config.RACES if r not in by_race)
    if silent:
        print("[import-polls] no usable polls (falls back to betting "
              "markets): %s" % ", ".join(silent))
    return path
